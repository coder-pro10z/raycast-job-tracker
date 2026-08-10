import openpyxl
import psycopg
from datetime import datetime
import os

# Grab connection string from environment variable or replace this string
DB_CONNECTION_STRING = os.environ.get("SUPABASE_CONNECTION_STRING", "your-supabase-connection-string")
EXCEL_FILE_PATH = '../public/Master_Job_Tracker.xlsx'

print(f"Connecting to Postgres: {DB_CONNECTION_STRING.split('@')[-1] if '@' in DB_CONNECTION_STRING else '...'} ")

try:
    conn = psycopg.connect(DB_CONNECTION_STRING)
    cur = conn.cursor()
except Exception as e:
    print(f"Failed to connect to database: {e}")
    print("Please set the SUPABASE_CONNECTION_STRING environment variable or edit the script directly.")
    exit(1)

print(f"Loading Excel file: {EXCEL_FILE_PATH}...")
try:
    wb = openpyxl.load_workbook(EXCEL_FILE_PATH, data_only=True)
except Exception as e:
    print(f"Failed to load Excel file: {e}")
    exit(1)

PRIORITY_MAP = {
    'High': 'High', '🔴 High': 'High',
    'Medium': 'Medium', 'Normal': 'Medium',
    'Low': 'Low'
}

seen = set()  # dedupe protection
inserted_count = 0

print("Beginning data migration...")

for sheet_name in wb.sheetnames:
    print(f"Processing sheet: {sheet_name}")
    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        continue
        
    header, data = rows[0], rows[1:]

    for row in data:
        r = dict(zip(header, row))
        company_name = r.get('Company Name')
        target_role = r.get('Target Role')
        
        # Skip empty rows
        if not company_name:
            continue
            
        key = (company_name, target_role)
        if key in seen:
            continue
            
        seen.add(key)
        
        # Safe extraction with defaults
        domain = r.get('Domain')
        if domain not in ('SDE / FullStack', 'Cloud / DevOps', 'Dual Domain'):
            domain = 'Dual Domain'
            
        work_mode = r.get('Work Mode')
        if work_mode not in ('Onsite', 'Hybrid', 'Remote'):
            work_mode = None
            
        status = r.get('Application Status')
        if status not in ('Not Started', 'Applied', 'In Review', 'Interview', 'Offer', 'Rejected', 'Withdrawn'):
            status = 'Not Started'
            
        priority_raw = str(r.get('Priority', 'Medium'))
        priority = PRIORITY_MAP.get(priority_raw, 'Medium')
        
        applied_date = r.get('Applied Date')
        if isinstance(applied_date, str):
            try:
                # Try to parse string dates if they aren't datetime objects
                applied_date = datetime.strptime(applied_date, "%Y-%m-%d").date()
            except ValueError:
                applied_date = None
        elif isinstance(applied_date, datetime):
            applied_date = applied_date.date()

        try:
            # 1. Insert into jobs table
            cur.execute("""
                INSERT INTO jobs (company_name, target_role, domain, location, work_mode,
                    application_link, priority, application_status, next_action, tech_stack,
                    career_page_link, applied_date, referral_needed, referral_contact_name,
                    hr_recruiter_name)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
                RETURNING id
            """, (
                company_name, target_role, domain, r.get('Location'), work_mode,
                r.get('Application Link'), priority, status, r.get('Next Action'), r.get('Tech Stack'),
                r.get('Career Page Link'), applied_date,
                (str(r.get('Referral Needed')).lower() == 'yes'), r.get('Referral Contact Name'), r.get('HR/Recruiter Name')
            ))
            
            job_id = cur.fetchone()[0]
            inserted_count += 1

            # 2. Insert Notes if present
            notes = r.get('Notes')
            if notes:
                cur.execute(
                    "INSERT INTO notes (job_id, content, note_type) VALUES (%s, %s, 'General')",
                    (job_id, str(notes))
                )
                
        except Exception as e:
            print(f"Error inserting row for {company_name}: {e}")
            conn.rollback()
            continue

print("Committing transaction...")
conn.commit()

print("=================================================")
print(f"✅ Migration Complete!")
print(f"Total Unique Companies Processed: {len(seen)}")
print(f"Total Jobs Successfully Inserted: {inserted_count}")
print("=================================================")

cur.close()
conn.close()
